from __future__ import annotations

from io import BytesIO

import openpyxl

from app.comptroller.export import COLUMNS, build_review_queue_workbook


def review_row(**overrides):
    row = {
        "review_month": "2026-08-01",
        "comptroller_business_name": "ACME HARDWARE",
        "comptroller_legal_name": "ACME HARDWARE LLC",
        "comptroller_taxpayer_id": "17512000001",
        "comptroller_location_number": "1",
        "comptroller_address": "100 MAIN ST",
        "comptroller_city": "LUBBOCK",
        "comptroller_state": "TX",
        "comptroller_zip": "79401",
        "comptroller_permit_start_date": "2010-01-01",
        "comptroller_permit_end_date": "2026-08-01",
        "comptroller_previous_status": "ACTIVE",
        "comptroller_current_status": "INACTIVE",
        "first_detected_at": "2026-08-12T00:00:00+00:00",
        "matched_account_number": None,
        "matched_owner_name": None,
        "match_confidence": "UNMATCHED",
        "match_score": 0.0,
        "match_ambiguous": False,
        "match_reason": "No RenditionPilot rendition records found for this district.",
        "workflow_status": "PENDING_REVIEW",
        "reviewer_notes": None,
        "reviewed_at": None,
    }
    row.update(overrides)
    return row


def load_workbook_rows(xlsx_bytes: bytes):
    workbook = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return rows


def test_workbook_has_header_and_one_row_per_review():
    reviews = [review_row(), review_row(comptroller_business_name="HOT TACO")]

    xlsx_bytes = build_review_queue_workbook(reviews, month_label="2026-08")
    rows = load_workbook_rows(xlsx_bytes)

    assert rows[0] == tuple(label for _, label in COLUMNS)
    assert len(rows) == 3  # header + 2 data rows
    business_name_col = [label for _, label in COLUMNS].index("Business / DBA")
    assert rows[1][business_name_col] == "ACME HARDWARE"
    assert rows[2][business_name_col] == "HOT TACO"


def test_workbook_handles_zero_reviews():
    xlsx_bytes = build_review_queue_workbook([], month_label="2026-08")
    rows = load_workbook_rows(xlsx_bytes)

    assert rows == [tuple(label for _, label in COLUMNS)]


def test_workbook_header_is_bold():
    xlsx_bytes = build_review_queue_workbook([review_row()], month_label="2026-08")
    workbook = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active

    assert all(cell.font.bold for cell in sheet[1])


def test_sheet_name_is_truncated_to_excels_limit():
    xlsx_bytes = build_review_queue_workbook([], month_label="2026-08-a-very-long-suffix-that-overflows")
    workbook = openpyxl.load_workbook(BytesIO(xlsx_bytes))

    assert len(workbook.sheetnames[0]) <= 31
